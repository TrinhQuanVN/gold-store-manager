import psycopg2
from openpyxl import load_workbook

# === Kết nối PostgreSQL ===
conn = psycopg2.connect(
    dbname="nextjsHQ",
    user="postgres",
    password="91Nguyenvantroi",
    host="localhost",
    port="5432"
)

header_id = 1

cur = conn.cursor()
# cur.execute("TRUNCATE TABLE \"TaxPayer\" RESTART IDENTITY CASCADE;")
# conn.commit()
# taxpayer = [["Tên người nộp thuế: Doanh nghiệp tư nhân vàng bạc đá quý Hồng Quân",
#              "0700756585",
#              "SN 91, đường Nguyễn Văn Trỗi, phường Phủ Lý, tỉnh Ninh Bình"],
#           ]
# for name, taxcode, address in taxpayer:
#     cur.execute(
#         """
#         INSERT INTO "TaxPayer" ("name", "taxCode", "address")
#         VALUES (%s, %s, %s)
#         """,
#         (name, taxcode, address),
#     )

# conn.commit()

# cur.execute("TRUNCATE TABLE \"ReportXNTGroup\" RESTART IDENTITY CASCADE;")
# conn.commit()
# groups = [["Vàng 9999 (24K)","1"],["Vàng trang sức 999 (23K)","2"],
#           ["Vàng trang sức 999 (23K)","3"],
#           ["Vàng trang sức 75% (18K)","4"],
#           ["Vàng trang sức 58,5% (14K)","5"],
#           ["Vàng trang sức 41,6% (10K)","6"],
#           ["Bạc 999","7"],]
# for name, stt in groups:
#     cur.execute(
#         """
#         INSERT INTO "ReportXNTGroup" ("name", "stt", "headerId")
#         VALUES (%s, %s, %s)
#         """,
#         (name, stt, header_id),
#     )

# conn.commit()

# === 1. RESET bảng ReportXNT ===
# cur.execute("TRUNCATE TABLE \"ReportXNT\" RESTART IDENTITY CASCADE;")
# conn.commit()
print("✅ Đã reset bảng ReportXNT.")

# === 2. Đọc file Excel ===
excel_path = r"C:\Users\Admin\Desktop\duLieuVangQ1_2025.xlsx"  # 👉 thay bằng đường dẫn thật
wb = load_workbook(excel_path)
ws = wb.active



# Bỏ qua dòng header nếu có
row_start = 2 if isinstance(ws["A1"].value, str) else 1

inserted_count = 0

for row in ws.iter_rows(min_row=row_start, values_only=True):
    if not row or not row[0]:  # Bỏ dòng rỗng
        continue
    (
        groupId, stt, id, name,
        tonDauKyQuantity, tonDauKyValue,
        nhapQuantity, nhapValue,
        xuatQuantity, xuatValue,
        tonCuoiKyQuantity, tonCuoiKyValue
    ) = row[:12]

    cur.execute("""
        INSERT INTO "ReportXNT" (
            "groupId", "stt","id", "name", "unit",
            "tonDauKyQuantity", "tonDauKyValue",
            "nhapQuantity", "nhapValue",
            "xuatQuantity", "xuatValue",
            "tonCuoiKyQuantity", "tonCuoiKyValue"
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        groupId, stt, id, str(name), "chỉ",
        tonDauKyQuantity or 0, tonDauKyValue or 0,
        nhapQuantity or 0, nhapValue or 0,
        xuatQuantity or 0, xuatValue or 0,
        tonCuoiKyQuantity or 0, tonCuoiKyValue or 0
    ))

    inserted_count += 1

conn.commit()
print(f"✅ Đã chèn {inserted_count} dòng từ Excel vào PostgreSQL.")

# === Đóng kết nối ===
cur.close()
conn.close()
